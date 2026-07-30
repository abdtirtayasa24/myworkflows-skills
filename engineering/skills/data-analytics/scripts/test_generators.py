from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent


class NotebookGeneratorTest(unittest.TestCase):
    def test_creates_structured_analysis_notebook_and_protects_existing_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "churn-analysis.ipynb"
            command = [
                sys.executable,
                str(SCRIPTS_DIR / "create_analysis_notebook.py"),
                "--title",
                "Churn Analysis",
                "--objective",
                "Identify accounts at risk of churn",
                "--target",
                "churned_within_30_days",
                "--metric",
                "PR-AUC",
                "--out",
                str(output),
            ]

            first_run = subprocess.run(command, capture_output=True, text=True, check=False)
            self.assertEqual(first_run.returncode, 0, first_run.stderr)

            notebook = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(notebook["nbformat"], 4)
            self.assertEqual(notebook["cells"][0]["cell_type"], "markdown")
            rendered = "".join(
                "".join(cell["source"])
                for cell in notebook["cells"]
                if cell["cell_type"] == "markdown"
            )
            for expected in (
                "# Churn Analysis",
                "Identify accounts at risk of churn",
                "churned_within_30_days",
                "PR-AUC",
                "## Data validation and cleaning",
                "## Exploratory data analysis",
                "## Modeling and evaluation",
                "## Synthesis and quality control",
            ):
                self.assertIn(expected, rendered)

            second_run = subprocess.run(command, capture_output=True, text=True, check=False)
            self.assertNotEqual(second_run.returncode, 0)
            self.assertIn("--force", second_run.stderr)

            forced_run = subprocess.run(
                [*command, "--force"], capture_output=True, text=True, check=False
            )
            self.assertEqual(forced_run.returncode, 0, forced_run.stderr)


class ExcelDashboardGeneratorTest(unittest.TestCase):
    def test_creates_styled_dashboard_workbook_with_formulas_and_chart(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "churn-dashboard.xlsx"
            command = [
                sys.executable,
                str(SCRIPTS_DIR / "create_excel_dashboard.py"),
                "--title",
                "Churn Dashboard",
                "--objective",
                "Monitor churn risk",
                "--target",
                "churned_within_30_days",
                "--metric",
                "Monthly churn rate",
                "--out",
                str(output),
            ]

            result = subprocess.run(command, capture_output=True, text=True, check=False)
            self.assertEqual(result.returncode, 0, result.stderr)

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["README", "Dashboard", "KPI Summary", "Clean Data"],
            )
            self.assertEqual(workbook["README"]["A1"].value, "Churn Dashboard")
            self.assertEqual(workbook["README"]["B5"].value, "Monthly churn rate")
            self.assertEqual(workbook["KPI Summary"]["D4"].value, '=IF(OR(B4="",C4=""),"",B4-C4)')
            self.assertEqual(workbook["KPI Summary"]["E3"].value, "Direction")
            self.assertEqual(workbook["KPI Summary"]["F3"].value, "Status")
            self.assertIn('E4="Higher is better"', workbook["KPI Summary"]["F4"].value)
            self.assertIn('E4="Lower is better"', workbook["KPI Summary"]["F4"].value)
            self.assertEqual(len(workbook["KPI Summary"].data_validations.dataValidation), 1)
            self.assertEqual(workbook["KPI Summary"].freeze_panes, "A4")
            self.assertEqual(workbook["Clean Data"].freeze_panes, "A2")
            self.assertEqual(len(workbook["Dashboard"]._charts), 1)
            self.assertTrue(workbook["Dashboard"].sheet_view.showGridLines is False)

            protected_run = subprocess.run(command, capture_output=True, text=True, check=False)
            self.assertNotEqual(protected_run.returncode, 0)
            self.assertIn("--force", protected_run.stderr)

    def test_escapes_formula_like_user_text(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "safe-dashboard.xlsx"
            command = [
                sys.executable,
                str(SCRIPTS_DIR / "create_excel_dashboard.py"),
                "--title",
                "Safe Dashboard",
                "--objective",
                '=HYPERLINK("https://example.invalid", "click")',
                "--target",
                "+2+2",
                "--metric",
                "@SUM(A1:A2)",
                "--out",
                str(output),
            ]

            result = subprocess.run(command, capture_output=True, text=True, check=False)
            self.assertEqual(result.returncode, 0, result.stderr)

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook["README"]["B3"].value,
                '\'=HYPERLINK("https://example.invalid", "click")',
            )
            self.assertEqual(workbook["README"]["B4"].value, "'+2+2")
            self.assertEqual(workbook["KPI Summary"]["A4"].value, "'@SUM(A1:A2)")


if __name__ == "__main__":
    unittest.main()
