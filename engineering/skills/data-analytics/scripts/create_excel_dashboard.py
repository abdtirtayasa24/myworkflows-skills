#!/usr/bin/env python3
"""Create a structured Excel dashboard skeleton for data analysis."""

from __future__ import annotations

import argparse
import re
from datetime import UTC, datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Run with: uv run --with openpyxl python "
        "scripts/create_excel_dashboard.py ..."
    ) from exc


NAVY = "17324D"
BLUE = "1F77B4"
LIGHT_BLUE = "D9EAF7"
LIGHT_GRAY = "EEF2F5"
WHITE = "FFFFFF"
GREEN = "D9EAD3"
RED = "F4CCCC"
DARK_TEXT = "22313F"
THIN_GRAY = Side(style="thin", color="D5DCE3")


def slugify(value: str) -> str:
    """Return a filesystem-friendly slug."""
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "dashboard"


def safe_excel_text(value: str) -> str:
    """Prevent user-controlled labels from being interpreted as formulas."""
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def style_title(sheet: Worksheet, title: str, end_column: str = "F") -> None:
    """Add a consistent title band to a worksheet."""
    sheet.merge_cells(f"A1:{end_column}1")
    cell = sheet["A1"]
    cell.value = safe_excel_text(title)
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.sheet_view.showGridLines = False


def style_header_row(sheet: Worksheet, row: int, start: int, end: int) -> None:
    """Apply table-header styling to a range of cells."""
    for column in range(start, end + 1):
        cell = sheet.cell(row=row, column=column)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN_GRAY)


def build_readme(
    workbook: Workbook,
    title: str,
    objective: str,
    target: str,
    metric: str,
    generated_at: str,
) -> None:
    """Create the workbook guide and analysis contract."""
    sheet = workbook.active
    sheet.title = "README"
    style_title(sheet, title)

    rows = [
        (3, "Objective", objective),
        (4, "Target / KPI", target),
        (5, "Primary metric", metric),
        (6, "Generated (UTC)", generated_at),
        (7, "Status", "Skeleton — replace placeholders and validate before distribution"),
    ]
    for row, label, value in rows:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=safe_excel_text(value))
        sheet.cell(row=row, column=1).font = Font(bold=True, color=DARK_TEXT)
        sheet.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    sheet["A9"] = "Workbook structure"
    sheet["A9"].font = Font(bold=True, color=WHITE)
    sheet["A9"].fill = PatternFill("solid", fgColor=BLUE)
    guidance = [
        ("Dashboard", "Stakeholder-facing KPI cards and chart; avoid raw-data details."),
        ("KPI Summary", "Enter reviewed KPI values, targets, definitions, and owners."),
        ("Clean Data", "Load analysis-ready records only; preserve source data elsewhere."),
    ]
    for row, (name, purpose) in enumerate(guidance, start=10):
        sheet.cell(row=row, column=1, value=name).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=purpose).alignment = Alignment(wrap_text=True)

    sheet["A14"] = "Quality-control checklist"
    sheet["A14"].font = Font(bold=True, color=WHITE)
    sheet["A14"].fill = PatternFill("solid", fgColor=BLUE)
    checks = [
        "☐ KPI definitions, units, time window, and denominators are confirmed.",
        "☐ Dashboard totals reconcile to Clean Data and the source.",
        "☐ Filters, formulas, charts, and refresh instructions are tested.",
        "☐ Limitations, uncertainty, and as-of date are visible to readers.",
        "☐ Sensitive data and hidden metadata have been reviewed.",
    ]
    for row, check in enumerate(checks, start=15):
        sheet.cell(row=row, column=1, value=check)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 72
    for column in ("C", "D", "E", "F"):
        sheet.column_dimensions[column].width = 14
    sheet.freeze_panes = "A3"


def build_kpi_summary(workbook: Workbook, title: str, metric: str) -> None:
    """Create the editable KPI table and status formulas."""
    sheet = workbook.create_sheet("KPI Summary")
    style_title(sheet, f"{title} — KPI Summary", end_column="G")

    headers = [
        "KPI",
        "Current",
        "Target",
        "Variance",
        "Direction",
        "Status",
        "Definition / Owner",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column, value=header)
    style_header_row(sheet, row=3, start=1, end=len(headers))

    labels = [metric, "Guardrail 1", "Guardrail 2", "Segment KPI 1", "Segment KPI 2"]
    labels.extend(f"KPI {number}" for number in range(6, 11))
    for row, label in enumerate(labels, start=4):
        sheet.cell(row=row, column=1, value=safe_excel_text(label))
        sheet.cell(row=row, column=4, value=f'=IF(OR(B{row}="",C{row}=""),"",B{row}-C{row})')
        sheet.cell(
            row=row,
            column=6,
            value=(
                f'=IF(OR(B{row}="",C{row}="",E{row}=""),"",'
                f'IF(E{row}="Higher is better",IF(B{row}>=C{row},"On track","Off track"),'
                f'IF(E{row}="Lower is better",IF(B{row}<=C{row},"On track","Off track"),'
                '"Check direction")))'
            ),
        )
        for column in range(1, 8):
            cell = sheet.cell(row=row, column=column)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in (2, 3, 5, 7):
            sheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        for column in (2, 3, 4):
            sheet.cell(row=row, column=column).number_format = "#,##0.00"

    direction_validation = DataValidation(
        type="list",
        formula1='"Higher is better,Lower is better"',
        allow_blank=True,
    )
    direction_validation.promptTitle = "Metric direction"
    direction_validation.prompt = "Choose how the KPI should be evaluated against its target."
    direction_validation.error = "Choose Higher is better or Lower is better."
    direction_validation.errorTitle = "Invalid direction"
    direction_validation.showErrorMessage = True
    sheet.add_data_validation(direction_validation)
    direction_validation.add("E4:E13")

    sheet.conditional_formatting.add(
        "F4:F13",
        FormulaRule(formula=['F4="On track"'], fill=PatternFill("solid", fgColor=GREEN)),
    )
    sheet.conditional_formatting.add(
        "F4:F13",
        FormulaRule(formula=['F4="Off track"'], fill=PatternFill("solid", fgColor=RED)),
    )
    widths = {"A": 28, "B": 14, "C": 14, "D": 14, "E": 18, "F": 16, "G": 45}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = "A3:G13"
    sheet.freeze_panes = "A4"


def build_dashboard(workbook: Workbook, title: str) -> None:
    """Create a stakeholder-facing dashboard with linked KPI cards and chart."""
    sheet = workbook.create_sheet("Dashboard", 1)
    style_title(sheet, title, end_column="H")
    sheet["A2"] = "Populate the KPI Summary sheet; this page updates from its formulas."
    sheet["A2"].font = Font(italic=True, color=DARK_TEXT)

    card_starts = [("A3", "A4", "A5", 4), ("E3", "E4", "E5", 5)]
    for title_cell, value_cell, status_cell, source_row in card_starts:
        title_column = sheet[title_cell].column
        end_column = title_column + 2
        for row in range(3, 6):
            for column in range(title_column, end_column + 1):
                sheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                sheet.cell(row=row, column=column).border = Border(
                    left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY
                )
        sheet.merge_cells(start_row=3, start_column=title_column, end_row=3, end_column=end_column)
        sheet.merge_cells(start_row=4, start_column=title_column, end_row=4, end_column=end_column)
        sheet.merge_cells(start_row=5, start_column=title_column, end_row=5, end_column=end_column)
        sheet[title_cell] = f"='KPI Summary'!A{source_row}"
        sheet[value_cell] = f"='KPI Summary'!B{source_row}"
        sheet[status_cell] = f"='KPI Summary'!F{source_row}"
        sheet[title_cell].font = Font(bold=True, color=DARK_TEXT)
        sheet[value_cell].font = Font(size=20, bold=True, color=BLUE)
        for cell_ref in (title_cell, value_cell, status_cell):
            sheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

    kpi_sheet = workbook["KPI Summary"]
    chart = BarChart()
    chart.title = "Current vs Target"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "KPI"
    chart.height = 8
    chart.width = 16
    chart.add_data(Reference(kpi_sheet, min_col=2, max_col=3, min_row=3, max_row=13), titles_from_data=True)
    chart.set_categories(Reference(kpi_sheet, min_col=1, min_row=4, max_row=13))
    sheet.add_chart(chart, "A8")

    for column in "ABCDEFGH":
        sheet.column_dimensions[column].width = 14
    sheet.freeze_panes = "A3"


def build_clean_data(workbook: Workbook) -> None:
    """Create an analysis-ready data sheet with a formatted starter table."""
    sheet = workbook.create_sheet("Clean Data")
    headers = ["record_id", "date", "segment", "target", "value", "notes"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)
    style_header_row(sheet, row=1, start=1, end=len(headers))

    for column in range(1, len(headers) + 1):
        sheet.cell(row=2, column=column, value=None)
    table = Table(displayName="CleanData", ref="A1:F2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F2"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 40
    sheet.sheet_view.showGridLines = False


def build_workbook(title: str, objective: str, target: str, metric: str) -> Workbook:
    """Build the complete dashboard workbook."""
    workbook = Workbook()
    workbook.properties.title = title
    workbook.properties.subject = "Data analytics dashboard skeleton"
    workbook.properties.creator = "data-analytics skill"
    generated_at = datetime.now(UTC).replace(microsecond=0).isoformat()

    build_readme(workbook, title, objective, target, metric, generated_at)
    build_kpi_summary(workbook, title, metric)
    build_dashboard(workbook, title)
    build_clean_data(workbook)
    return workbook


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--title", required=True, help="Human-readable dashboard title.")
    parser.add_argument("--objective", default="[Confirm the business objective]", help="Decision or objective.")
    parser.add_argument("--target", default="[Confirm the target or KPI]", help="Target variable or KPI.")
    parser.add_argument("--metric", default="[Confirm the success metric]", help="Primary success metric.")
    parser.add_argument("--out", type=Path, help="Output .xlsx path.")
    parser.add_argument("--force", action="store_true", help="Overwrite an existing output file.")
    return parser.parse_args()


def main() -> None:
    """Generate the requested dashboard workbook."""
    args = parse_args()
    output = args.out or Path("output/data-analytics") / f"{slugify(args.title)}.xlsx"
    if output.suffix.lower() != ".xlsx":
        raise SystemExit(f"Output must use the .xlsx extension: {output}")
    if output.exists() and not args.force:
        raise SystemExit(f"Refusing to overwrite existing file without --force: {output}")

    workbook = build_workbook(args.title, args.objective, args.target, args.metric)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(f"Wrote Excel dashboard: {output.resolve()}")


if __name__ == "__main__":
    main()
